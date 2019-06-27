import os, PyPDF2, re
from PyPDF2 import PdfFileReader
from os import listdir
import os.path
from os.path import isfile, join
import warnings
#import tempfile
import PIL.Image
from wand.image import Image
from wand.color import Color
import requests
import glob
import skimage
import pytesseract
import cv2
import numpy
import string
import shutil
from word2number import w2n
import PythonMagick
from openpyxl import Workbook
#import base64
import openpyxl

#Prerequesite:  Ghostscript -    'pip install ghostscript' and https://www.ghostscript.com/download/gsdnld.html
#               Imagemagick - https://imagemagick.org/script/download.php
#               Tesseract-OCR - https://github.com/UB-Mannheim/tesseract/wiki



def get_url():
    #ENTER URL IN QUOTATIONS BELOW
    URL = 'https://ptsv2.com/t/z4f72-1561445479/post'
    return URL

def check_connection():
    url=get_url()
    timeout=5
    try:
        _ = requests.get(url, timeout=timeout)
        return True
    except requests.ConnectionError:
        return False    

def ocr(filename):      
    text = pytesseract.image_to_string(PIL.Image.open(filename))  
    return text


def split_pdf_pages(root_directory, extract_to_folder): 
    for root, dirs, files in os.walk(root_directory):
        for filename in files:
            basename, extension = os.path.splitext(filename)
            if extension == ".pdf":
                fullpath = root + "\\" + basename + extension
                opened_pdf = PyPDF2.PdfFileReader(open(fullpath,"rb"),False, None, True)
                for i in range(opened_pdf.numPages):
                    output = PyPDF2.PdfFileWriter()
                    output.addPage(opened_pdf.getPage(i))
                    with open(extract_to_folder + "\\" + basename + "-%s.pdf" % i, "wb") as output_pdf:
                        output.write(output_pdf)





def convert_pdf(filename, output_path, resolution):

        all_pages = Image(filename=filename, resolution=resolution)
        for i, page in enumerate(all_pages.sequence):
            with Image(page) as img:
                img.format = 'jpg'
                img.background_color = Color('white')
                img.alpha_channel = 'remove'

                image_filename = os.path.splitext(os.path.basename(filename))[0]
                image_filename = '{}-{}.jpg'.format(image_filename, i)
                image_filename = os.path.join(output_path, image_filename)

                img.save(filename=image_filename)

def files_list(directory):
    xlist = [f for f in listdir(directory) if isfile(join(directory, f))]
    return xlist



def img_rot(x, name):
    for i in range(x):
        img_name = name[i]
        img_path = 'CREATE TEMPLATE/img/' + img_name
        im = skimage.io.imread(img_path)
        newdata = pytesseract.image_to_osd(im, nice=1)
        rotation=(re.search('(?<=Rotate: )\d+', newdata).group(0))
        rotation = int(rotation)
        if(rotation !=0):
            unrotated_img = img_path
            imgx = cv2.imread(unrotated_img) 
            (rows, cols) = imgx.shape[:2]  
            M = cv2.getRotationMatrix2D((cols / 2, rows / 2), rotation, 1) 
            res = cv2.warpAffine(imgx, M, (cols, rows))
            rot_img_path = 'CREATE TEMPLATE/rimg/' + img_name
            cv2.imwrite(rot_img_path, res)
        else:
            rot_img_path = 'CREATE TEMPLATE/rimg/' + img_name
            shutil.copy(img_path,rot_img_path)


def remove_underlines(path):
    img = cv2.imread(path, 0)
    img = cv2.bitwise_not(img)  

    kernel_clean = numpy.ones((2,2),numpy.uint8)
    cleaned = cv2.erode(img, kernel_clean, iterations=1)

    kernel_line = numpy.ones((1, 5), numpy.uint8)  
    clean_lines = cv2.erode(cleaned, kernel_line, iterations=6)
    clean_lines = cv2.dilate(clean_lines, kernel_line, iterations=6)

    cleaned_img_without_lines = cleaned - clean_lines
    cleaned_img_without_lines = cv2.bitwise_not(cleaned_img_without_lines)

    cv2.imwrite(path, cleaned_img_without_lines)


def check_duplicate(listx, directory):
    img_sig = []
    for f in listx:
        file_dir = directory + f
        imgg = PythonMagick.Image(file_dir)
        sig = imgg.signature()
        if(sig not in img_sig[:]):
            img_sig.append(sig)
        else:
            os.remove(file_dir)

def max_v(sheetx):
    a=0
    for row in sheetx.iter_rows(min_row=6,min_col=2, max_col=2, max_row=1000006):
        for cell in row:
            if(cell.value !=None):
                a+=1
            else:
                return a
            
def matching(list1, str1):
    match = [m for m in list1 if m in str1]
    if(match):
        return str(match[0])
    else:
        return None

def pn(value):
    if(value>0):
        return value-1
    else:
        return value

def extract(raw_excel_str1,ocr_string):
    ocr_str = ocr_string.split('\n')
    if((raw_excel_str1.count(';') == 0) and (raw_excel_str1.count(':') == 0)):
        a1 = raw_excel_str1.split(',')
        row = pn(int(a1[0]))
        #print(row)
        col = pn(int(a1[1]))
        #print(col)
        raw1 = ocr_str[row]
        #print(raw1)
        raw2 = raw1.split()
        #print(raw2)
        raw3 = str(raw2[col])
        return raw3
    elif(raw_excel_str1.count(';')>0):
        item = raw_excel_str1.split(';')
        raw4 = ''
        for i in range(len(item)):
            a1 = item.split(',')                    
            row = pn(int(a1[i]))
            col = pn(int(a1[i+1]))
            raw1 = ocr_str[row]
            raw2 = raw1.split()
            raw3 = str(raw2[col])
            raw4+=raw3
        return raw4
    elif(raw_excel_str1.count(':')== 1):
        item = raw_excel_str1.split(':')
        a1 = item[0]
        #print(a1)
        a2 = item[1]
        #print(a2)
        a1x = a1.split(',')
        #print(a1x)
        a2x = a2.split(',')
        #print(a2x)
        if(a1x[0] == a2x[0]):
            ocr_lines = ocr_str
            #print(ocr_lines)
            ocr_row_str = ocr_lines[pn(int(a1x[0]))]
            ocr_row = ocr_row_str.split()
            raw = ocr_row[pn(int(a1x[1])):pn(int(a2x[1]))]
            raw.append(str(ocr_row[pn(int(a2x[1]))]))
            rawx = ' '.join(raw)
            
            return rawx
        else:
            raw = "None"
            return raw

def amount_words_extract(raw_excel_str,ocr_str):
    ocr_lines = ocr_str.split('\n')
    row_col = raw_excel_str.split(',')
    row = pn(int(row_col[0]))
    col = pn(int(row_col[1]))
    amount_line = ocr_lines[row]
    amount_line = amount_line[col:]
    if('only' in amount_line.lower()):
        amount = amount_line[:amount_line.lower().index('only')-1]
        return amount
    else:
        amount = "None"
        return amount

def amount_space_extract(raw_excel_str,ocr_str):
    ocr_lines = ocr_str.split('\n')
    row_col = raw_excel_str.split(',')
    row = pn(int(row_col[0]))
    col = pn(int(row_col[1]))
    amount_l = ocr_lines[row]
    amount_line = amount_l.split()
    amount_line = amount_line[col:]    
    amount = ''.join(amount_line)        
    return amount
    
        
def names():
    wb = openpyxl.load_workbook('CREATE TEMPLATE\Templates.xlsx')
    sheet = wb.get_sheet_by_name('sheet1')
    name_list = []
    for row in sheet.iter_rows(min_row = 6, min_col = 2, max_row=max_v(sheet)+5, max_col = 2):
        for cell in row:
            name_list.append(cell.value)
            
    return name_list

    

def inv_decode(rot_img_name,vname,ocr_full_img):
    path_of_image = 'CREATE TEMPLATE/rimg/' + rot_img_name
    path_of_template = 'CREATE TEMPLATE\Template storage\\' + vname + ' ts.jpg'
    with open(path_of_image, "rb") as im:
        with open(path_of_template, "rb") as tm:
            wb = openpyxl.load_workbook('CREATE TEMPLATE\Templates.xlsx')
            sheet = wb.get_sheet_by_name('sheet1')
            name_list = []
            for row in sheet.iter_rows(min_row = 6, min_col = 2, max_row=max_v(sheet)+5, max_col = 2):
                for cell in row:
                    name_list.append(cell.value)
            for i in range(len(name_list)):
                name_list[i] = name_list[i].lower()
            name = matching(name_list, ocr_full_img)
            name_index = int(name_list.index(name)+6)
            x=int(name_index)
            inv = cv2.imread(path_of_image)
            template = cv2.imread(path_of_template)
            ht = inv.shape[0]
            wt = inv.shape[1]
            d = (wt,ht)
            template_resize = cv2.resize(template, d, interpolation = cv2.INTER_AREA)
            out = cv2.add(template_resize,inv)
            out = cv2.blur(out,(4,4))
            cv2.imwrite('CREATE TEMPLATE\cache 2\\output.jpg', out)
            fimg = 'CREATE TEMPLATE\cache 2\\output.jpg'
            if(sheet.cell(row = x, column = 15).value == 1):
                remove_underlines(fimg)
            ocr_inv = ocr(fimg)
            un_list = ["'",'"', ':', ',', ';', '?', '/','<','.','>','|','[',']','{','}','(',')','*','^','!','&','%','$','#','@']
            ocr_lines = ocr_inv.split('\n')
            for i in range(len(ocr_lines)):
                line = ocr_lines[i].strip()
                if(len(line) == 1):
                    if(line[0] in un_list):
                        line = line.replace(line[0],"",1).strip()
                if(len(line)>1):
                    if(line[0] in un_list):
                        line = line.replace(line[0],"",1).strip()
                    if(line[-1] in un_list):
                        line = line[:-1].strip()
                ocr_lines[i] = line
            ocr_lines = [line for line in ocr_lines if line.strip() != ''] 
            ocr_inv = '\n'.join(ocr_lines)
            #print(ocr_inv)
            
            
            if(sheet.cell(row = name_index, column = 1).value == (name_index-6)):
                code = x-6
                v_name = sheet.cell(row = x, column = 2).value
                if(sheet.cell(row = x, column = 5).value == 1):
                    raw_inv_no = sheet.cell(row = x, column = 3).value
                    inv_no = extract(raw_inv_no,ocr_inv)
                    raw_inv_date = sheet.cell(row = x, column = 4).value
                    inv_date = extract(raw_inv_date,ocr_inv)
                              
                elif(sheet.cell(row = x, column = 5).value == 0):
                    raw_data = sheet.cell(row = x, column = 3).value
                    inv_no = extract(raw_data,ocr_inv)
                    inv_date = "None"
                    
                else:
                    raw_data = sheet.cell(row = x, column = 3).value
                    separator = sheet.cell(row = x, column = 5).value
                    raw_data_ = extract(raw_data,ocr_inv)
                    inv_no = raw_data_.split(separator,1)[0]
                    inv_date = raw_data_.split(separator,1)[1]
                   
                    
                if(sheet.cell(row = x, column = 9).value == 1):
                    if(sheet.cell(row = x, column = 6).value == 0):
                        if(sheet.cell(row = x, column = 8).value == 1):                       
                            raw_amount = sheet.cell(row = x, column = 7).value
                            inv_amount = amount_space_extract(raw_amount,ocr_inv)
                        elif(sheet.cell(row = x, column = 8).value == 0):
                            raw_amount = sheet.cell(row = x, column = 7).value
                            inv_amount = extract(raw_amount,ocr_inv)
                    elif(sheet.cell(row = x, column = 6).value == 1):
                        raw_amount = sheet.cell(row = x, column = 7).value
                        amount = amount_words_extract(raw_amount,ocr_inv)
                        inv_amount = w2n.word_to_num(amount)
                        if(amount == None):
                            inv_amount = "None"
                        else:
                            pass
                    else:
                        print("Wrong data in Templates.xlsx , row =" , x , "col = 6")
                        ux = input()
                    raw_cur = sheet.cell(row = x, column = 10).value
                    inv_currency = extract(raw_cur,ocr_inv)
                elif(sheet.cell(row = x,column = 9).value == 0):
                    if(sheet.cell(row = x, column = 6).value == 0):
                        if(sheet.cell(row = x, column = 8).value == 1):
                            raw_amount = sheet.cell(row = x, column = 7).value
                            inv_am = amount_space_extract(raw_amount,ocr_inv)
                            inv_amount = inv_am[3:]
                            inv_currency = inv_am[:3]
                        elif(sheet.cell(row = x, column = 8).value == 0):                            
                            raw_amount = sheet.cell(row = x, column = 7).value
                            inv_am = extract(raw_amount,ocr_inv)
                            inv_amount = inv_am[3:]
                            inv_currency = inv_am[:3]
                    elif(sheet.cell(row = x, column = 6).value == 1):
                        raw_amount = sheet.cell(row = x, column = 7).value
                        inv_am = amount_words_extract(raw_amount,ocr_inv)
                        #print(inv_am)
                        amount = inv_am[3:]
                        inv_currency = inv_am[:3]
                        #print(amount)
                        inv_amount = w2n.word_to_num(amount)
                        if(amount == None):
                            inv_amount = "None"
                        else:
                            pass
                    else:
                        print("Wrong data in Templates.xlsx  [row =" , x , "col = 6]")
                        ux = input()
                else:
                    sep = sheet.cell(row = x,column = 9).value
                    raw_datax = sheet.cell(row = x, column = 7).value
                    raw_datax_ = extract(raw_datax,ocr_inv)
                    cur = raw_datax.split(sep,1)[0]
                    amount = raw_datax.split(sep,1)[1]
                    if(sheet.cell(row = x, column = 6).value == 0):
                        if(' ' in amount):
                            amount = amount.replace(' ','')
                        inv_amount = amount
                        inv_currency = cur
                    elif(sheet.cell(row = x, column = 6).value == 1):
                        amount_line = amount
                        if('only' in amount_line.lower()):
                            amount = amount_line[:amount_line.lower().index('only')-1]
                        inv_amount = w2n.word_to_num(amount)
                        inv_currency = cur
                    else:
                        print("Wrong data in Templates.xlsx  [row =" , x , "col = 6]")
                        ux = input()
                    
                if(sheet.cell(row = x, column = 11).value == 1):
                    if(sheet.cell(row = x, column = 12).value == 1): 
                        raw_rs = sheet.cell(row = x, column = 13).value
                        inv_rs = extract(raw_rs, ocr_inv)
                        raw_rs_date = sheet.cell(row = x,column = 14).value
                        inv_rs_date = extract(raw_rs_date,ocr_inv)
                    elif(sheet.cell(row = x, column = 12).value == 0):
                        inv_rs = "None"
                        inv_rs_date = "None"
                    else:
                        raw_rs = sheet.cell(row = x, column = 13).value
                        sep_rs = sheet.cell(row = x, column = 12).value
                        raw_rs_ = extract(raw_rs,ocr_inv)
                        inv_rs = raw_rs_.split(sep_rs,1)[0]
                        inv_rs_date = raw_rs.split(sep_rs,1)[1]
                        
                        
                else:
                    inv_rs = "None"
                    inv_rs_date = "None"
                    
                    
            else:
                print("\nERROR: code!=list index")
                ux = input()
    return code,v_name,inv_no,inv_date,inv_amount,inv_currency,inv_rs,inv_rs_date

            
if __name__ == "__main__":
       
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\\tesseract.exe"

    warnings.filterwarnings("ignore")

    useless_2 = glob.glob('CREATE TEMPLATE/cache/*')
    for f in useless_2:
        os.remove(f)
        
    useless_3 = glob.glob('CREATE TEMPLATE/img/*')
    for f in useless_3:
        os.remove(f)
        
    useless_4 = glob.glob('CREATE TEMPLATE/spdf/*')
    for f in useless_4:
        os.remove(f)

    useless_5 = glob.glob('CREATE TEMPLATE/rimg/*')
    for f in useless_5:
        os.remove(f)
        
    # parameter variables  
    root_dir = r"____INVOICE____(PDF)"
    extract_to = r"CREATE TEMPLATE/spdf"    
    split_pdf_pages(root_dir, extract_to)
    split_list = files_list('CREATE TEMPLATE\spdf')

    for i in range(len(split_list)):
        F = split_list[i]
        Fname = 'CREATE TEMPLATE\spdf\\' + F
        convert_pdf(Fname, 'CREATE TEMPLATE\img' ,300)
        
    img_list = files_list('CREATE TEMPLATE\img')
    img_rot(len(img_list),img_list)
    rot_img_list = files_list('CREATE TEMPLATE/rimg')
    check_duplicate(rot_img_list, 'CREATE TEMPLATE/rimg/')
    rot_img_list = files_list('CREATE TEMPLATE/rimg')

    ocr_out = []
    for i in range(len(rot_img_list)):
        F_OCR = rot_img_list[i]
        FILE_OCR = 'CREATE TEMPLATE/rimg/' + F_OCR
        output_of_ocr = ocr(FILE_OCR)
        ocr_out.append(output_of_ocr)
        #print(ocr_out[i])

    v_name_list = names()
    for i in range(len(v_name_list)):
        if(v_name_list[i] != None):
            v_name_list[i] = v_name_list[i].lower()
    #print(v_name_list)
        

    inv_count = 0
    vendor_excel_refno = []
    invoice_number = []
    invoice_date = []
    invoice_amount = []
    invoice_currency = []
    vendor_name = []
    invoice_rs_no = []
    invoice_rs_date = []

    po_vcode = []
    po_invparty = []
    po_off_vess = []
    po_place_port = []
    po_eta = []
    po_number = []
    po_date = []
    po_country = []
    po_etd = []
    po_currency = []
    po_file = []
    po_shipno = []

    n= "None"

    for i in range(len(rot_img_list)):
        info = ocr_out[i]
        length = len(info)
        if((('sci contact person' in info.lower()) == True)):
            po_file.append(rot_img_list[i])
            po_line = []
            for line in info.split('\n'):
                po_line.append(line)
            po_counter = 0
            data = []
            no_of_lines = len(po_line)
            po = re.search(r"(\d{9,10}/\d{2}/\d{2}/\d{4})", info)
            if(po == None):
                po_name = '' + rot_img_list[i]
                img = cv2.imread(po_name,0)
                
                img = cv2.bitwise_not(img)  

                kernel_clean = numpy.ones((2,2),numpy.uint8)
                cleaned = cv2.erode(img, kernel_clean, iterations=1)

                kernel_line = numpy.ones((1, 5), numpy.uint8)  
                clean_lines = cv2.erode(cleaned, kernel_line, iterations=6)
                clean_lines = cv2.dilate(clean_lines, kernel_line, iterations=6)

                cleaned_img_without_lines = cleaned - clean_lines
                cleaned_img_without_lines = cv2.bitwise_not(cleaned_img_without_lines)

                clean_po = 'CREATE TEMPLATE\clean img cache\\' +rot_img_list[i]
                cv2.imwrite(clean_po,cleaned_img_without_lines)
                po_ocr = ocr(clean_po)
                po_x = re.search(r"(\d{9,10}/\d{2}/\d{2}/\d{4})", po_ocr)
                data = po_x.group(1).split('/')
                po_number.append(data[0])
                name = ''
                name=name + data[1]+'/'+ data[2] + '/' + data [3]
                po_date.append(name)
                useless_files_x = glob.glob('CREATE TEMPLATE\clean img cache\*')
                for f in useless_files_x:
                    os.remove(f)
            else:
                data = po.group(1).split('/')
                po_number.append(data[0])
                name = ''
                name=name + data[1]+'/'+ data[2] + '/' + data [3]
                po_date.append(name)
            

            if(('vendor code' in info.lower()) != True):
                po_vcode.append(n)
            if(('invoicing party' in info.lower()) != True):
                po_invparty.append(n)
            if(('sci office/vessel' in info.lower()) != True):
                po_off_vess.append(n)
            if(('place/port' in info.lower()) != True):
                po_place_port.append(n)
            if(('eta :' in info.lower()) != True):
                po_eta.append(n)
            if(('country' in info.lower()) != True):
                po_country.append(n)
            if((('etd :' in info.lower()) != True) and (('etd:' in info.lower()) != True)):
                po_etd.append(n)
            if(('pr no ' in info.lower()) != True):
                po_currency.append(n)
            if('danaos pr no:-' not in info.lower()):
                po_shipno.append(n)
                
            for d in range(no_of_lines):
                item = po_line[d]
                if(('vendor code' in item.lower()) == True):
                    data=item.split()
                    po_vcode.append(data[3])
                if(('invoicing party' in item.lower()) == True):
                    data = item.split()
                    po_invparty.append(data[-1])
                if(('sci office/vessel' in item.lower()) == True):
                    if('country' not in item.lower()):
                        data=item[item.index(':')+2:]                        
                    else:
                        data = item[item.index(':')+2:item.lower().index('country')-1]
                    po_off_vess.append(data)
                if(('place/port' in item.lower()) == True):
                    data = item.split()
                    flag1 = []
                    portname = ''
                    countryname = ''
                    for kk in range(len(data)):
                        if(data[kk] == ":"):
                            flag1.append(kk)
                        elif(data[kk].lower() == "country"):
                            flag2 = kk
                        elif(data[kk].lower() == "business"):
                            flag2 = kk
                    if((("country" in item.lower()) !=True) and ('business' not in item.lower())):
                        flag2 = len(data)
                    
                        
                    for kk in range((flag1[0]+1),flag2):
                        portname=portname + data[kk]+" "
                    po_place_port.append(portname)
                    
                    
                if(('eta :' in item.lower()) == True):
                    data=item.split()
                    etaflag = []
                    for kk in range(len(data)):
                        if(data[kk] == ":"):
                            etaflag.append(kk)
                    po_eta.append(data[etaflag[0]+1])
                    
                if(('country :' in item.lower()) == True):
                    if(item.count(":") == 1):
                        data=item[item.index(':')+2:]
                    elif(item.count(":") == 2):
                        data = item[item.lower().index('country')+10:]
                        
                    po_country.append(data)
                if((('etd :' in item.lower()) == True) or (('etd:' in item.lower()) == True)):
                    data = item.split()
                    po_etd.append(data[-1])
                if(('pr no ' in item.lower()) == True):
                    data = item.split()
                    po_currency.append(data[-1])
                if('danaos pr no:-' in item.lower()):
                    raw = item.split()
                    data = raw[2]
                    po_shipno.append(data[4:8])
             
            po_counter+=1

        elif(("sci contact person" not in info.lower()) and (("invoice" in info.lower()) or ("bill" in info.lower()))):
            info = info.lower()
            v_name = matching(v_name_list,info)
            a,b,c,d,e,f,g,h = inv_decode(rot_img_list[i],v_name,info)
            vendor_excel_refno.append(a)
            vendor_name.append(b)
            invoice_number.append(c)
            invoice_date.append(d)
            invoice_amount.append(e)
            invoice_currency.append(f)
            invoice_rs_no.append(g)
            invoice_rs_date.append(h)
            inv_count+=1
            
    
           
        
    for i in range(len(po_file)):
        if(po_currency[i]== "None"):
            file_name = po_file[i]
            file_n = "____INVOICE____(PDF)\\" + file_name[:-8] + ".pdf"
            pgnum = PdfFileReader(open(file_n, "rb")).getNumPages()
            image_name = po_file[i]
            for j in range(pgnum):
                image_name = image_name[:-7] + str(j) + "-0.jpg"
                index = rot_img_list.index(image_name)
                data = ocr_out[index]
                if(('including tax in' in data.lower()) == True):
                    datax = data[data.lower().index("including tax in")+17: data.lower().index("including tax in")+20]
                else:
                    continue
            po_currency[i] = datax
        if(po_shipno[i] == "None"):
            for d in range(len(po_line)):
                item = po_line[d]
                if('no:-' in item.lower()):
                    raw = item[item.lower().index('no:-')+4:8]
                    po_shipno[i] = raw


    if(inv_count>0):
        for i in range(inv_count):
            print("\nVENDOR NAME:\t", vendor_name[i], "\nINVOICE NO:\t", invoice_number[i], "\nINVOICE DATE:\t", invoice_date[i], "\nAMOUNT:  \t", invoice_amount[i])
            print("CURRENCY:\t", invoice_currency[i], "\nRS NUMBER:\t", invoice_rs_no[i], "\nRS DATE:  \t", invoice_rs_date[i])

    if(len(po_file)>0):
        print("\n\n\nPO\n")
        for i in range(len(po_file)):
            print("\n\n\nPO FILE:",po_file[i],"\n\nPO number:\t",po_number[i],"\nPO date:\t",po_date[i],"\nVendor Code:\t", po_vcode[i],"\nInvoice Party:\t", po_invparty[i],"\nVessel Code:\t", po_shipno[i], "\nVessel Name:\t",po_off_vess[i],"\nPort:\t\t",po_place_port[i],"\nETA:\t\t",po_eta[i],"\nPO country:\t",po_country[i],"\nPO etd:\t\t",po_etd[i],"\nPO currency:\t",po_currency[i])


    inv_info = []
    po_info = []

    if(inv_count>0):
        for i in range(inv_count):
            i_data = [vendor_name[i],invoice_number[i],invoice_date[i],invoice_amount[i],invoice_currency[i],invoice_rs_no[i],invoice_rs_date[i]]
            inv_info.append(i_data)

    if(len(po_file)>0):
        for i in range(len(po_file)):
            p_data = [po_number[i],po_date[i],po_vcode[i],po_invparty[i],po_off_vess[i],po_place_port[i],po_eta[i],po_country[i],po_etd[i],po_currency[i]]
            po_info.append(p_data)


    url = get_url()
    internet = check_connection()
    
    if(internet == True):   
        if(len(po_file) == 0):
            if(inv_count == 1):
                payload = inv_info[0] 
                r = requests.post(url, params = None,json = payload)
            elif(inv_count>1):
                payload = inv_info[0]
                for i in range(1,inv_count):
                    payload = payload + inv_info[i]
                r = requests.post(url,params = None, json = payload)
        elif((len(po_file) == 1) and (inv_count == 1)):
            payload = inv_info[0] + po_info[0]
            r = requests.post(url, params = None,json = payload)
        elif((len(po_file) == 1) and (inv_count == 0)):
            payload = po_info[0]
            r = requests.post(url, params = None,json = payload)
    else:
        print("\n\n\t.................ERROR: NO CONNECTION.................")
        uxx = input()
        
                


        
            
            
            

